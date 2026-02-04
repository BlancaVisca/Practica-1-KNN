import math
from openpyxl import Workbook
import tkinter as tk
from tkinter import filedialog, messagebox


datos_trained = None
clases_trained = None
datos_test = None
clases_test = None
datos_new = None
k_seleccionado = None

#Aqui se definen los arreglos que utiliza el GUI


#Formula de la distancia euclidiana, pensada para repetirse len(p1) veces
def distancia_euclidiana(p1, p2):
    suma = 0
    for i in range(len(p1)):
        suma += (p1[i] - p2[i]) ** 2
    return math.sqrt(suma)

#Funcion para subir archivos y separar datos y clases
def subir_archivo(nombre_archivo,tiene_clase=True):
    datos = []
    clases = []
    with open(nombre_archivo, "r") as archivo:
        for linea in archivo:
            linea = linea.strip()
            if linea == "":
                continue

            partes = linea.split(",")

            atributos = [
                float(partes[0]),
                float(partes[1]),
                float(partes[2]),
                float(partes[3])
            ]
            datos.append(atributos)
            if tiene_clase:
                clases.append(partes[4])

    if tiene_clase:
        return datos, clases
    else:
        return datos

#Funcion principal del algoritmo KNN
#Utilizamos todos los valores de trained, y un punto en especifico de test para calcular el KNN
#Se ordena de menor a mayor la distancia y se devuelven los k vecinos mas cercanos
def knn(trained, clases, punto_prueba, k):
    distancias = []

    for i in range(len(trained)):
        d = distancia_euclidiana(punto_prueba, trained[i])
        distancias.append((d, clases[i]))

    distancias.sort(key=lambda x: x[0])
    vecinos = distancias[:k]
    return vecinos


#Funcion para calcular la exactitud del modelo
#Se utilizan todos los valores de trained y test, ademas de un valor K para determinar la exactitud total del modelo
#Se utiliza un punto de test para predecir su clase, y se compara con la clase real, si acierta se suma 1 a correctos
def exactitud(datos_trained, clases_trained, datos_test, clases_test, k):
    correctos = 0
    

    for i in range(len(datos_test)):
        prediccion = knn(datos_trained, clases_trained, datos_test[i], k)
        mayoria = mayoria_clase(prediccion)
        if mayoria == clases_test[i]:
            correctos += 1

    exactitud = correctos / len(datos_test)
    return exactitud

#Funcion para determinar la clase mayoritaria entre los k vecinos mas cercanos
def mayoria_clase(prediccion):
    clase_vecinos = []
    for j in range(len(prediccion)):
        clase_vecinos.append(prediccion[j][1])
    mayoria = max(clase_vecinos, key=clase_vecinos.count)
    return mayoria

#Carga de archivos
datos_trained, clases_trained= subir_archivo("DataTrained-iris.data",tiene_clase=True)
datos_test, clases_test= subir_archivo("TestData-iris.data",tiene_clase=True)
datos_new= subir_archivo("NewData-iris.data",tiene_clase=False)


#Pruebas de exactitud con diferentes valores de K
#Creacion de archivos Excel con los resultados

#print("Exactitud del modelo" , exactitud(datos_trained, clases_trained, datos_test, clases_test, 3) , "\n")

#print("Exactitud del modelo", exactitud(datos_trained, clases_trained, datos_test, clases_test, 5), "\n")

#print("Exactitud del modelo" , exactitud(datos_trained, clases_trained, datos_test, clases_test, 7), "\n")


#wb = Workbook()
#hoja = wb.active

#hoja.append(["Objeto prueba", "Clase", "K3", "K5", "K7"])
#for i in range(len(datos_test)):
    #objeto = ", ".join(map(str, datos_test[i]))
    #hoja.append([objeto, clases_test[i],mayoria_clase(knn(datos_trained, clases_trained, datos_test[i], 3)),mayoria_clase(knn(datos_trained, clases_trained, datos_test[i], 5)),mayoria_clase(knn(datos_trained, clases_trained, datos_test[i], 7))])

#wb.save("Set-Prueba-357.xlsx")



#print("Clasificación de nuevos datos (sin clase) con k=3:\n")
#wb2 = Workbook()
#hoja2 = wb2.active

#hoja2.append(["Objeto prueba", "Clase asignada"])
#for i in range(len(datos_new)):
    #vecinos = knn(datos_trained, clases_trained, datos_new[i], 3)
    #clase_predicha = mayoria_clase(vecinos)

    #objeto2 = ", ".join(map(str, datos_new[i]))
    #hoja2.append([objeto2, clase_predicha])
#wb2.save("New-Data-Best-K.xlsx")



#Ventana GUI

def cargar_entrenamiento():
    global datos_trained, clases_trained
    archivo = filedialog.askopenfilename(
        title="Cargar datos de entrenamiento",
        filetypes=[("Data files", "*.data *.txt")]
    )
    if archivo:
        datos_trained, clases_trained = subir_archivo(archivo, True)
        messagebox.showinfo("OK", "Datos de entrenamiento cargados")

def cargar_prueba():
    global datos_test, clases_test
    archivo = filedialog.askopenfilename(
        title="Cargar datos de prueba",
        filetypes=[("Data files", "*.data *.txt")]
    )
    if archivo:
        datos_test, clases_test = subir_archivo(archivo, True)
        messagebox.showinfo("OK", "Datos de prueba cargados")

def cargar_nuevos():
    global datos_new
    archivo = filedialog.askopenfilename(
        title="Cargar nuevos datos",
        filetypes=[("Data files", "*.data *.txt")]
    )
    if archivo:
        datos_new = subir_archivo(archivo, False)
        messagebox.showinfo("OK", "Nuevos datos cargados")

def ejecutar_knn_gui():
    if datos_trained is None or datos_new is None:
        messagebox.showerror("Error", "Carga entrenamiento y nuevos datos")
        return

    try:
        k = int(entry_k.get())
    except:
        messagebox.showerror("Error", "K debe ser un número entero")
        return

    texto_resultados.delete("1.0", tk.END)

    if datos_test is not None:
        acc = exactitud(datos_trained, clases_trained, datos_test, clases_test, k)
        texto_resultados.insert(
            tk.END,
            f"Exactitud del modelo con k={k}: {acc}\n\n"
        )

    texto_resultados.insert(tk.END, "Clasificación de nuevos datos:\n")

    wb = Workbook()
    hoja = wb.active
    hoja.title = "New Best K"
    hoja.append(["Objeto", "Clase asignada"])

    for i in range(len(datos_new)):
        vecinos = knn(datos_trained, clases_trained, datos_new[i], k)
        pred = mayoria_clase(vecinos)
        objeto = ", ".join(map(str, datos_new[i]))

        texto_resultados.insert(
            tk.END,
            f"Objeto {i+1}: {objeto} → {pred}\n"
        )

        hoja.append([objeto, pred])

    wb.save("New-Best-K.xlsx")
    messagebox.showinfo("OK", "Archivo New-Best-K.xlsx creado")

# =========================
# GUI
# =========================

ventana = tk.Tk()
ventana.title("Práctica 1: Clasificador K-NN")
ventana.geometry("800x500")
ventana.configure(bg="#f2f2f2")

# ----- TÍTULO -----
tk.Label(
    ventana,
    text="Clasificador K-NN (Iris Dataset)",
    font=("Arial", 18, "bold"),
    bg="#f2f2f2"
).pack(pady=10)

# ----- FRAME CARGA DE ARCHIVOS -----
frame_archivos = tk.LabelFrame(
    ventana,
    text="Carga de Archivos",
    font=("Arial", 11, "bold"),
    padx=10,
    pady=10,
    bg="#f2f2f2"
)
frame_archivos.columnconfigure(0, weight=1)
frame_archivos.pack(fill="x", padx=20, pady=40)

tk.Button(
    frame_archivos,
    text="1. Cargar Entrenamiento",
    width=30,
    command=cargar_entrenamiento
).grid(row=0, column=0, pady=5)

tk.Button(
    frame_archivos,
    text="2. Cargar Prueba (Test)",
    width=30,
    command=cargar_prueba
).grid(row=1, column=0, pady=5)

tk.Button(
    frame_archivos,
    text="3. Cargar Nuevos Datos",
    width=30,
    command=cargar_nuevos
).grid(row=2, column=0, pady=5)




# ----- FRAME K -----
frame_k = tk.Frame(ventana, bg="#f2f2f2")
frame_k.pack(pady=5)

tk.Label(
    frame_k,
    text="Valor de K:",
    font=("Arial", 11),
    bg="#f2f2f2"
).pack(side="left", padx=5)

entry_k = tk.Entry(frame_k, width=10, justify="center")
entry_k.pack(side="left")

# ----- BOTÓN EJECUTAR -----
tk.Button(
    ventana,
    text="EJECUTAR KNN Y CLASIFICAR",
    bg="#2196F3",
    fg="white",
    font=("Arial", 11, "bold"),
    width=35,
    command=ejecutar_knn_gui
).pack(pady=10)

# ----- ÁREA DE RESULTADOS -----
frame_resultados = tk.LabelFrame(
    ventana,
    text="Resultados",
    font=("Arial", 11, "bold"),
    padx=10,
    pady=10,
    bg="#f2f2f2"
)
frame_resultados.pack(fill="both", expand=True, padx=20, pady=10)

texto_resultados = tk.Text(
    frame_resultados,
    height=12,
    font=("Consolas", 10)
)
texto_resultados.pack(fill="both", expand=True)

texto_resultados.insert(
    tk.END,
    "Bienvenido.\nCargue los archivos y seleccione el valor de K para comenzar.\n"
)

ventana.mainloop()




